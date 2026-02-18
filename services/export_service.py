# -*- coding: utf-8 -*-
"""
Export service for CSV, Excel, and GeoJSON exports.
Implements FR-D-15 Data Export requirements.
"""

import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from models.building import Building
from repositories.database import Database
from repositories.building_repository import BuildingRepository
from services.export.export_strategy import CSVExportStrategy
from utils.logger import get_logger

logger = get_logger(__name__)


class ExportService:
    """Service for exporting data to various formats."""

    def __init__(self, db: Database):
        self.db = db
        self.building_repo = BuildingRepository(db)
        self._csv_strategy = CSVExportStrategy()

    def _write_csv_file(
        self,
        data: List[Dict[str, Any]],
        file_path: Path,
        columns: Optional[List[str]] = None
    ) -> bool:
        """
        Internal helper to write CSV using CSVExportStrategy.

        Maintains backwards compatibility with existing behavior.
        """
        return self._csv_strategy.export(
            data=data,
            file_path=str(file_path),
            columns=columns,
            encoding='utf-8-sig'
        )

    def export_buildings_csv(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None,
        columns: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Export buildings to CSV file.

        Args:
            file_path: Output file path
            filters: Optional filters (neighborhood, type, status)
            columns: Optional list of columns to include

        Returns:
            Export summary dict
        """
        # Default columns
        if columns is None:
            columns = [
                "building_id",
                "neighborhood_name",
                "neighborhood_name_ar",
                "building_type",
                "building_status",
                "number_of_units",
                "number_of_apartments",
                "number_of_shops",
                "number_of_floors",
                "latitude",
                "longitude",
            ]

        # Get buildings with filters
        buildings = self.building_repo.search(
            neighborhood_code=filters.get("neighborhood_code") if filters else None,
            building_type=filters.get("building_type") if filters else None,
            building_status=filters.get("building_status") if filters else None,
            limit=10000  # Max export
        )

        # Convert to list of dicts
        data = [building.to_dict() for building in buildings]

        # Write CSV using strategy
        self._write_csv_file(data, file_path, columns)

        logger.info(f"Exported {len(buildings)} buildings to {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(buildings),
            "format": "csv",
            "exported_at": datetime.now().isoformat()
        }

    def export_buildings_excel(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export buildings to Excel file.

        Args:
            file_path: Output file path
            filters: Optional filters

        Returns:
            Export summary dict
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed")
            raise ImportError("openpyxl is required for Excel export")

        # Get buildings
        buildings = self.building_repo.search(
            neighborhood_code=filters.get("neighborhood_code") if filters else None,
            building_type=filters.get("building_type") if filters else None,
            building_status=filters.get("building_status") if filters else None,
            limit=10000
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Buildings"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072BC", end_color="0072BC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = [
            "Building ID", "Neighborhood", "Neighborhood (AR)",
            "Type", "Status", "Units", "Apartments", "Shops",
            "Floors", "Latitude", "Longitude"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_num, building in enumerate(buildings, 2):
            data = [
                building.building_id,
                building.neighborhood_name,
                building.neighborhood_name_ar,
                building.building_type_display,
                building.building_status_display,
                building.number_of_units,
                building.number_of_apartments,
                building.number_of_shops,
                building.number_of_floors,
                building.latitude,
                building.longitude,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        column_widths = [25, 15, 15, 12, 12, 8, 10, 8, 8, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Save
        wb.save(file_path)
        logger.info(f"Exported {len(buildings)} buildings to {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(buildings),
            "format": "xlsx",
            "exported_at": datetime.now().isoformat()
        }

    def get_export_preview(
        self,
        filters: Optional[Dict[str, Any]] = None,
        limit: int = 10
    ) -> List[Dict[str, Any]]:
        """Get preview of data to be exported."""
        buildings = self.building_repo.search(
            neighborhood_code=filters.get("neighborhood_code") if filters else None,
            building_type=filters.get("building_type") if filters else None,
            building_status=filters.get("building_status") if filters else None,
            limit=limit
        )

        return [b.to_dict() for b in buildings]

    def get_export_count(self, filters: Optional[Dict[str, Any]] = None) -> int:
        """Get count of records to be exported."""
        return self.building_repo.count(filters)

    # ==================== GeoJSON Export (FR-D-15.3) ====================

    def export_buildings_geojson(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export buildings to GeoJSON file (FR-D-15.3).

        Args:
            file_path: Output file path
            filters: Optional filters

        Returns:
            Export summary dict
        """
        # Get buildings
        buildings = self.building_repo.search(
            neighborhood_code=filters.get("neighborhood_code") if filters else None,
            building_type=filters.get("building_type") if filters else None,
            building_status=filters.get("building_status") if filters else None,
            limit=10000
        )

        # Build GeoJSON FeatureCollection
        features = []
        for building in buildings:
            if building.latitude and building.longitude:
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [building.longitude, building.latitude]
                    },
                    "properties": {
                        "building_id": building.building_id,
                        "building_uuid": building.building_uuid,
                        "neighborhood_name": building.neighborhood_name,
                        "neighborhood_name_ar": building.neighborhood_name_ar,
                        "building_type": building.building_type,
                        "building_status": building.building_status,
                        "number_of_units": building.number_of_units,
                        "number_of_floors": building.number_of_floors,
                        "full_address": building.full_address,
                        "full_address_ar": building.full_address_ar,
                        "legacy_stdm_id": building.legacy_stdm_id,
                    }
                }

                # Add polygon geometry if available
                if building.geo_location:
                    try:
                        geo_data = json.loads(building.geo_location)
                        if geo_data.get("type") == "Polygon":
                            feature["geometry"] = geo_data
                    except json.JSONDecodeError:
                        pass

                features.append(feature)

        geojson = {
            "type": "FeatureCollection",
            "name": "TRRCMS_Buildings",
            "crs": {
                "type": "name",
                "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}
            },
            "features": features,
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "record_count": len(features),
                "source": "UN-Habitat TRRCMS"
            }
        }

        # Write to file
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(geojson, f, ensure_ascii=False, indent=2)

        logger.info(f"Exported {len(features)} buildings to GeoJSON: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(features),
            "format": "geojson",
            "exported_at": datetime.now().isoformat()
        }

    def export_units_geojson(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export units to GeoJSON file (FR-D-15.3).

        Args:
            file_path: Output file path
            filters: Optional filters

        Returns:
            Export summary dict
        """
        cursor = self.db.cursor()

        query = """
            SELECT u.*, b.latitude, b.longitude, b.geo_location,
                   b.neighborhood_name, b.neighborhood_name_ar
            FROM property_units u
            JOIN buildings b ON u.building_id = b.building_id
            WHERE 1=1
        """
        params = []

        if filters:
            if filters.get("building_id"):
                query += " AND u.building_id = ?"
                params.append(filters["building_id"])
            if filters.get("unit_type"):
                query += " AND u.unit_type = ?"
                params.append(filters["unit_type"])

        query += " LIMIT 10000"
        cursor.execute(query, params)

        features = []
        for row in cursor.fetchall():
            data = dict(row)
            if data.get("latitude") and data.get("longitude"):
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [data["longitude"], data["latitude"]]
                    },
                    "properties": {
                        "unit_uuid": data.get("unit_uuid"),
                        "unit_id": data.get("unit_id"),
                        "building_id": data.get("building_id"),
                        "unit_type": data.get("unit_type"),
                        "unit_number": data.get("unit_number"),
                        "floor_number": data.get("floor_number"),
                        "apartment_status": data.get("apartment_status"),
                        "area_sqm": data.get("area_sqm"),
                        "neighborhood_name": data.get("neighborhood_name"),
                        "neighborhood_name_ar": data.get("neighborhood_name_ar"),
                        "legacy_stdm_id": data.get("legacy_stdm_id"),
                    }
                }
                features.append(feature)

        geojson = {
            "type": "FeatureCollection",
            "name": "TRRCMS_Units",
            "crs": {
                "type": "name",
                "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}
            },
            "features": features,
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "record_count": len(features),
                "source": "UN-Habitat TRRCMS"
            }
        }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(geojson, f, ensure_ascii=False, indent=2)

        logger.info(f"Exported {len(features)} units to GeoJSON: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(features),
            "format": "geojson",
            "exported_at": datetime.now().isoformat()
        }

    def export_claims_geojson(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export claims to GeoJSON file (FR-D-15.3).

        Args:
            file_path: Output file path
            filters: Optional filters

        Returns:
            Export summary dict
        """
        cursor = self.db.cursor()

        query = """
            SELECT c.*, u.unit_id, b.latitude, b.longitude, b.geo_location,
                   b.neighborhood_name, b.neighborhood_name_ar, b.building_id
            FROM claims c
            LEFT JOIN property_units u ON c.unit_id = u.unit_id
            LEFT JOIN buildings b ON u.building_id = b.building_id
            WHERE 1=1
        """
        params = []

        if filters:
            if filters.get("case_status"):
                query += " AND c.case_status = ?"
                params.append(filters["case_status"])
            if filters.get("claim_type"):
                query += " AND c.claim_type = ?"
                params.append(filters["claim_type"])

        query += " LIMIT 10000"
        cursor.execute(query, params)

        features = []
        for row in cursor.fetchall():
            data = dict(row)
            if data.get("latitude") and data.get("longitude"):
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [data["longitude"], data["latitude"]]
                    },
                    "properties": {
                        "claim_id": data.get("claim_id"),
                        "claim_uuid": data.get("claim_uuid"),
                        "case_number": data.get("case_number"),
                        "case_status": data.get("case_status"),
                        "claim_type": data.get("claim_type"),
                        "priority": data.get("priority"),
                        "unit_id": data.get("unit_id"),
                        "building_id": data.get("building_id"),
                        "neighborhood_name": data.get("neighborhood_name"),
                        "neighborhood_name_ar": data.get("neighborhood_name_ar"),
                        "submission_date": data.get("submission_date"),
                        "has_conflict": data.get("has_conflict"),
                        "legacy_stdm_id": data.get("legacy_stdm_id"),
                    }
                }
                features.append(feature)

        geojson = {
            "type": "FeatureCollection",
            "name": "TRRCMS_Claims",
            "crs": {
                "type": "name",
                "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}
            },
            "features": features,
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "record_count": len(features),
                "source": "UN-Habitat TRRCMS"
            }
        }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(geojson, f, ensure_ascii=False, indent=2)

        logger.info(f"Exported {len(features)} claims to GeoJSON: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(features),
            "format": "geojson",
            "exported_at": datetime.now().isoformat()
        }

    # ==================== Claims Export (FR-D-15.2) ====================

    def export_claims_csv(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export claims to CSV file."""
        cursor = self.db.cursor()

        query = "SELECT * FROM claims WHERE 1=1"
        params = []

        if filters:
            if filters.get("case_status"):
                query += " AND case_status = ?"
                params.append(filters["case_status"])

        query += " ORDER BY created_at DESC LIMIT 10000"
        cursor.execute(query, params)

        columns = [
            "claim_id", "case_number", "case_status", "claim_type",
            "priority", "unit_id", "person_ids", "submission_date",
            "decision_date", "assigned_to", "has_conflict", "notes"
        ]

        rows = cursor.fetchall()

        # Convert to list of dicts
        data = [dict(row) for row in rows]

        # Write CSV using strategy
        self._write_csv_file(data, file_path, columns)

        logger.info(f"Exported {len(rows)} claims to CSV: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(rows),
            "format": "csv",
            "exported_at": datetime.now().isoformat()
        }

    def export_claims_excel(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export claims to Excel file (FR-D-15.2)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        cursor = self.db.cursor()

        query = "SELECT * FROM claims WHERE 1=1"
        params = []

        if filters:
            if filters.get("case_status"):
                query += " AND case_status = ?"
                params.append(filters["case_status"])

        query += " ORDER BY created_at DESC LIMIT 10000"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Claims"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072BC", end_color="0072BC", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Headers
        headers = [
            "Claim ID", "Case Number", "Status", "Type", "Priority",
            "Unit ID", "Submission Date", "Decision Date", "Assigned To",
            "Conflict", "Notes"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Data
        for row_num, row in enumerate(rows, 2):
            data = dict(row)
            values = [
                data.get("claim_id"), data.get("case_number"),
                data.get("case_status"), data.get("claim_type"),
                data.get("priority"), data.get("unit_id"),
                data.get("submission_date"), data.get("decision_date"),
                data.get("assigned_to"), "Yes" if data.get("has_conflict") else "No",
                data.get("notes")
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border

        wb.save(file_path)
        logger.info(f"Exported {len(rows)} claims to Excel: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(rows),
            "format": "xlsx",
            "exported_at": datetime.now().isoformat()
        }

    # ==================== Persons Export ====================

    def export_persons_csv(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export persons to CSV file."""
        cursor = self.db.cursor()

        query = "SELECT * FROM persons WHERE 1=1"
        params = []

        if filters:
            if filters.get("nationality"):
                query += " AND nationality = ?"
                params.append(filters["nationality"])

        query += " ORDER BY last_name, first_name LIMIT 10000"
        cursor.execute(query, params)

        columns = [
            "person_id", "first_name", "first_name_ar", "father_name", "father_name_ar",
            "last_name", "last_name_ar", "gender", "year_of_birth", "nationality",
            "national_id", "phone_number", "mobile_number", "email", "address"
        ]

        rows = cursor.fetchall()

        # Convert to list of dicts
        data = [dict(row) for row in rows]

        # Write CSV using strategy
        self._write_csv_file(data, file_path, columns)

        logger.info(f"Exported {len(rows)} persons to CSV: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(rows),
            "format": "csv",
            "exported_at": datetime.now().isoformat()
        }

    def export_units_csv(
        self,
        file_path: Path,
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export units to CSV file."""
        cursor = self.db.cursor()

        query = "SELECT * FROM property_units WHERE 1=1"
        params = []

        if filters:
            if filters.get("building_id"):
                query += " AND building_id = ?"
                params.append(filters["building_id"])
            if filters.get("unit_type"):
                query += " AND unit_type = ?"
                params.append(filters["unit_type"])

        query += " ORDER BY building_id, unit_number LIMIT 10000"
        cursor.execute(query, params)

        columns = [
            "unit_uuid", "unit_id", "building_id", "unit_type", "unit_number",
            "floor_number", "apartment_number", "apartment_status",
            "property_description", "area_sqm"
        ]

        rows = cursor.fetchall()

        # Convert to list of dicts
        data = [dict(row) for row in rows]

        # Write CSV using strategy
        self._write_csv_file(data, file_path, columns)

        logger.info(f"Exported {len(rows)} units to CSV: {file_path}")

        return {
            "file_path": str(file_path),
            "record_count": len(rows),
            "format": "csv",
            "exported_at": datetime.now().isoformat()
        }
